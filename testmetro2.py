import gc
import matplotlib.pyplot as plt
import os
import pandas as pd
from datetime import datetime
import xlwt
from sklearn.preprocessing import MinMaxScaler
import csv
import openpyxl
def get_base_features(df_):
    df = df_.copy()

    # base time
    # df['day'] = df['time'].apply(lambda x: int(x[8:10]))
    df['date']=df['date'].apply(transfer_datetoday)
    df['week']=df['date'].apply(transfer_realweek)
    # df['week'] = pd.to_datetime(df['time']).dt.dayofweek + 1
    df['weekend'] = df['date'].apply(transfer_realweekend)
    df['hour'] = df['time'].apply(transfer_timetohour)
    df['minute'] = df['time'].apply(transfer_timetominute)

    # count,sum
    result = df.groupby(['stationid', 'week', 'weekend', 'date', 'hour', 'minute']).status.agg(
        ['count', 'sum']).reset_index()

    # # nunique
    # tmp = df.groupby(['stationid'])['cardnum'].nunique().reset_index(name='cardnums_station')
    # result = result.merge(tmp, on=['stationid'], how='left')
    # tmp = df.groupby(['stationid', 'hour'])['cardnum'].nunique().reset_index(name='cardnums_station_hour')
    # result = result.merge(tmp, on=['stationid', 'hour'], how='left')
    # tmp = df.groupby(['stationid', 'hour', 'minute'])['cardnum'].nunique().reset_index(name='cardnums_station_hout_minute')
    # result = result.merge(tmp, on=['stationid', 'hour', 'minute'], how='left')

    # in,out
    result['outNums'] = result['sum']
    result['inNums'] = result['count'] - result['sum']

    # result['day_since_first']=result['date'].apply(transfer_date)
    result.fillna(0, inplace=True)
    # del result['sum'], result['count']

    return result

def transfer_datetoday(d):
    if len(d)==8:
        a=int(d[-1])
    else:
        a=int(d[-2:])
    return a

def transfer_realweek(d):
    if d in [4,5,6,11,12,18,19,25,26]:
        return 0
    else:
        return 1

def transfer_realweekend(d):
    if d in [4,5,6,11,12,18,19,25,26]:
        return 1
    else:
        return 0

def transfer_Timetoday(d):
    timelist=d.split(' ')
    if len(timelist[0]) == 8:
        a=int(timelist[0][-1])
        return a
    else:
        a=int(timelist[0][-2:])
        return a

def transfer_Timetohour(d):
    timelist = d.split(' ')
    if len(timelist[1]) == 7:
        a=int(timelist[1][0])
        return a
    else:
        a=int(timelist[1][0:2])
        return a

def transfer_Timetominute(d):
    timelist = d.split(' ')
    if len(timelist[1]) == 7:
        a=int(int(timelist[1][2])*10)
        return a
    else:
        a=int(int(timelist[1][3])*10)
        return a

def transfer_timetohour(d):
    # timelist = d.split(' ')
    if len(d) == 7:
        a=int(d[0])
        return a
    else:
        a=int(d[0:2])
        return a

def transfer_timetominute(d):
    # timelist = d.split(' ')
    if len(d) == 7:
        a=int(int(d[2])*10)
        return a
    else:
        a=int(int(d[3])*10)
        return a
def new3():
    test=pd.read_csv('_10mins/testsubmit.csv')
    # test['date']=test['startTime'].apply(transfer_Timetoday)
    test['hour']=test['startTime'].apply(transfer_Timetohour)
    test['minute']=test['startTime'].apply(transfer_Timetominute)

    # test['week']=test['date'].apply(transfer_realweek)
    # test['weekend']=test['date'].apply(transfer_realweekend)
    test = test.drop(['startTime','inNums','outNums'], axis=1)
    a=test.copy()

    date = [datetime.strftime(x, '%Y%m%d') for x in list(pd.date_range('20150427', '20150430'))]
    for i in date:
        bfn='_10mins/new2/'+str(i)+'new2.csv'
        test_sh=pd.read_csv(bfn)
        test_sh=test_sh.drop(['Unnamed: 0'], axis=1)
        # print(test_sh.date.unique())
        # all_columns = [f for f in test_sh.columns ]
        # X_data = test_sh[all_columns]
        test=test.merge(test_sh, on=['stationid','hour','minute'], how='left')
        test['date']=test_sh['date'][0]
        test['week'] = test_sh['week'][0]
        test['weekend'] = test_sh['weekend'][0]
        test.fillna(0, inplace=True)
        fn = str(i) + 'new3.csv'
        path = os.path.join("_10mins/new3", fn)
        test.to_csv(path)
        test=a

def new4():
    date = [datetime.strftime(x, '%Y%m%d') for x in list(pd.date_range('20150427', '20150430'))]
    for i in date:
        bfn = '_10mins/new3/' + str(i) + 'new3.csv'
        test_sh = pd.read_csv(bfn)
        test_sh.drop(test_sh[test_sh.hour<5].index,inplace=True)
        test_sh.drop(test_sh[test_sh.hour>=23].index, inplace=True)
        fn = str(i) + 'new4.csv'
        path = os.path.join("_10mins/new4", fn)
        test_sh.to_csv(path)


def draw():
    # 折线图
    x = []  # 点的横坐标
    for i in range(5,23):
        for j in range(6):
            time=str(i)+':'+str(j)
            x.append(time)
    test=pd.read_csv('_10mins/test.csv')
    test.drop(test[test.hour < 5].index, inplace=True)
    test.drop(test[test.hour >= 23].index, inplace=True)
    test=test[test.stationid==10]
    k1=test['outNums']
    result=pd.read_csv('_10mins/result4.csv')
    result=result[result.stationid==10]
    k2=result['outNums']/2
    # result2 = pd.read_csv('_10mins/sub_model9.csv')
    # result2 = result2[result2.stationid == 1]
    # k3 = result2['inNums'] * 2
    plt.plot(x, k1, 's-', color='r', label="REAL")  # s-:方形
    plt.plot(x, k2, 'o-', color='g', label="PREDICT")  # o-:圆形
    # plt.plot(x, k3, 's-', color='b', label="REAL")  # s-:方形
    plt.xlabel("time")  # 横坐标名字
    plt.ylabel("outNums")  # 纵坐标名字
    plt.legend(loc="best")  # 图例
    plt.show()

def tran_inNumsint(d):
    return int(d*2.5+0.5)
def tran_outNumssint(d):
    return int(d*32+0.5)
def tran_result():
    result=pd.read_csv('_10mins/sub_model12.csv')
    result['inNums']=result['inNums'].apply(tran_inNumsint)
    result['outNums']=result['outNums'].apply(tran_outNumssint)
    result.to_csv('_10mins/result4.csv')

def tran_jing(d):
    strs=d.split(' ')
    a = strs[0]
    return a
def tran_wei(d):
    strs=d.split(' ')
    a=strs[1]
    return a
def tran_name(d):
    strs=d.split(' ')
    a = strs[2].replace("'",'')
    return a
def tran_name2(d):
    strs=d.replace('"','')
    strs = strs.replace(',', '')
    return strs

def tran_flowdetail(d1,d2):
    print(d1,d2)
def tran_place():
    jingwei=pd.read_csv('_10mins/jingweileft.csv',encoding='utf-8')
    all_columns = [f for f in jingwei.columns if f in ['线路', '站点名', '"站点名地铁站名",','id']]
    jingwei_left=jingwei[all_columns]
    jingright=pd.read_csv('_10mins/jingweiright.csv',encoding='utf-8')

    all_columns2 = [f for f in jingright.columns if f in ['latitude', 'longitude', '"站点名地铁站名",']]
    jingwei_right = jingright[all_columns2]
    print(jingwei_right)
    test = jingwei_left.merge(jingwei_right, on=['"站点名地铁站名",'], how='left')
    test.to_csv('_10mins/jingwei.csv')
def tran_flow():
    # result=pd.read_csv('_10mins/result4_1.csv',encoding='utf-8')
    test = pd.read_csv('_10mins/test.csv')
    test.drop(test[test.hour < 5].index, inplace=True)
    test.drop(test[test.hour >= 23].index, inplace=True)
    print(test.columns)
    id=pd.read_csv('_10mins/idshunxu.csv',encoding='utf-8')
    idtest = id.merge(test, on=['stationid'], how='left')
    # idtest.drop(['Unnamed: 0','Unnamed: 0.1'], axis=1, inplace=True)
    idtest.drop(['Unnamed: 0'], axis=1, inplace=True)
    idtest['flow']=idtest['inNums']-idtest['outNums']
    idtest.to_csv('_10mins/idtest2.csv',encoding='utf-8')

def read_id():
    start=[21,43,65,77,88,96,145,163,181,195,210,240,280,310]
    startindex=[]
    data=openpyxl.load_workbook('_10mins/idtest.xlsx')
    flow=pd.read_csv('_10mins/idtest.csv')
    flowlist=flow['flow'].values
    print(flowlist)
    table=data['idtest']
    flag=0
    for i in range(2,33806):
        if table.cell(i,2).value in start:
            flag=1
        else:
            flag=0
        if flag == 1:
            table.cell(i,8).value=flowlist[i-2]
        else:
            table.cell(i,8).value=flowlist[i-2]+int(table.cell(i-108,8).value)
    data.save(filename="_10mins/idtestpy.xlsx")
    # table = data['Sheet3']
    # allidlist=[]
    # for i in range(2,315):
    #     id=[]
    #     id.append(table.cell(i,3).value)
    #     allidlist.append(id)
    # print(allidlist)
    # col=['stationid']
    # with open('_10mins/idshunxu.csv', 'w', encoding='utf-8', newline='') as fp:
    #     # 写
    #     writer = csv.writer(fp)
    #     # 设置第一行标题头
    #     writer.writerow(col)
    #     # 将数据写入
    #     writer.writerows(allidlist)

def modify_date(d):
    strlist=d.split(" ")
    strs='2022/7/15 '
    strs=strs+strlist[1]
    print(strs)
    return strs

def minmax():
    yongdu=pd.read_csv('_10mins/subway_congestionresult.csv')
    value=yongdu['value'].values
    value=value.reshape(-1,1)
    transfer = MinMaxScaler(feature_range=[0, 10])  # feature_range设定归一化值得范围，minmax归一化最大值一定是右端点，最小值左端点
    data_tran = transfer.fit_transform(value)
    data_tran=data_tran.reshape(1,-1)
    print(data_tran[0])
    yongdu['value0-10']=data_tran[0]
    yongdu.to_csv('_10mins/subwayyongdu.csv')
dictlist=[{65: 0, 70: 65, 66: 70, 72: 66, 67: 72, 61: 67, 62: 61, 60: 62, 68: 60, 69: 68, 64: 69, 71: 64, 63: 71}, {77: 0, 76: 77, 74: 76, 78: 74, 80: 78, 73: 80, 75: 73, 79: 75}, {88: 0, 84: 88, 92: 84, 83: 92, 82: 83, 91: 82, 85: 91, 86: 85, 90: 86, 87: 90, 81: 87, 89: 81}, {96: 0, 119: 96, 93: 119, 97: 93, 101: 97, 113: 101, 98: 113, 105: 98, 115: 105, 109: 115, 118: 109, 120: 118, 108: 120, 99: 108, 102: 99, 107: 102, 95: 107, 100: 95, 117: 100, 111: 117, 94: 111, 110: 94, 103: 110, 104: 103, 114: 104, 112: 114, 106: 112, 116: 106}, {145: 0, 128: 145, 127: 128, 142: 127, 121: 142, 144: 121, 135: 144, 148: 135, 130: 148, 132: 130, 138: 132, 137: 138, 136: 137, 124: 136, 140: 124, 141: 140, 134: 141, 147: 134, 131: 147, 125: 131, 143: 125, 123: 143, 129: 123, 122: 129, 133: 122, 146: 133, 126: 146, 139: 126}, {163: 0, 164: 163, 161: 164, 152: 161, 170: 152, 157: 170, 169: 157, 176: 169, 160: 176, 151: 160, 175: 151, 177: 175, 162: 177, 149: 162, 155: 149, 156: 155, 153: 156, 154: 153, 158: 154, 171: 158, 174: 171, 167: 174, 173: 167, 165: 173, 150: 165, 172: 150, 168: 172, 159: 168, 166: 159}, {181: 0, 183: 181, 178: 183, 192: 178, 187: 192, 190: 187, 186: 190, 182: 186, 191: 182, 185: 191, 184: 185, 179: 184, 180: 179, 188: 180, 189: 188, 193: 189}, {195: 0, 202: 195, 203: 202, 194: 203, 198: 194, 196: 198, 199: 196, 197: 199, 201: 197, 200: 201}, {210: 0, 224: 210, 213: 224, 225: 213, 230: 225, 227: 230, 208: 227, 216: 208, 226: 216, 205: 226, 215: 205, 229: 215, 206: 229, 204: 206, 220: 204, 228: 220, 221: 228, 217: 221, 222: 217, 218: 222, 211: 218, 209: 211, 212: 209, 214: 212, 223: 214, 219: 223, 207: 219}, {240: 0, 246: 240, 237: 246, 241: 237, 256: 241, 257: 256, 258: 257, 239: 258, 245: 239, 236: 245, 260: 236, 232: 260, 242: 232, 231: 242, 259: 231, 261: 259, 243: 261, 254: 243, 235: 254, 255: 235, 234: 255, 233: 234, 252: 233, 249: 252, 251: 249, 238: 251, 244: 238, 250: 244, 247: 250, 248: 247, 253: 248}, {280: 0, 271: 280, 269: 271, 276: 269, 273: 276, 272: 273, 287: 272, 263: 287, 288: 263, 289: 288, 284: 289, 274: 284, 270: 274, 264: 270, 279: 264, 277: 279, 290: 277, 283: 290, 265: 283, 278: 265, 282: 278, 262: 282, 268: 262, 267: 268, 286: 267, 266: 286, 285: 266, 275: 285, 281: 275}]
idlist=[[65, 70, 66, 72, 67, 61, 62, 60, 68, 69, 64, 71, 63], [77, 76, 74, 78, 80, 73, 75, 79], [88, 84, 92, 83, 82, 91, 85, 86, 90, 87, 81, 89], [96, 119, 93, 97, 101, 113, 98, 105, 115, 109, 118, 120, 108, 99, 102, 107, 95, 100, 117, 111, 94, 110, 103, 104, 114, 112, 106, 116], [145, 128, 127, 142, 121, 144, 135, 148, 130, 132, 138, 137, 136, 124, 140, 141, 134, 147, 131, 125, 143, 123, 129, 122, 133, 146, 126, 139], [163, 164, 161, 152, 170, 157, 169, 176, 160, 151, 175, 177, 162, 149, 155, 156, 153, 154, 158, 171, 174, 167, 173, 165, 150, 172, 168, 159, 166], [181, 183, 178, 192, 187, 190, 186, 182, 191, 185, 184, 179, 180, 188, 189, 193], [195, 202, 203, 194, 198, 196, 199, 197, 201, 200], [210, 224, 213, 225, 230, 227, 208, 216, 226, 205, 215, 229, 206, 204, 220, 228, 221, 217, 222, 218, 211, 209, 212, 214, 223, 219, 207], [240, 246, 237, 241, 256, 257, 258, 239, 245, 236, 260, 232, 242, 231, 259, 261, 243, 254, 235, 255, 234, 233, 252, 249, 251, 238, 244, 250, 247, 248, 253], [280, 271, 269, 276, 273, 272, 287, 263, 288, 289, 284, 274, 270, 264, 279, 277, 290, 283, 265, 278, 282, 262, 268, 267, 286, 266, 285, 275, 281]]
startindex=[2, 2918, 6374, 7778, 8642, 9938, 12962, 15986, 19118, 20846, 21926, 24842, 28190, 31322]
minmax()
# draw()
# tran_flow()
# real=pd.rea
# d_csv('_10mins/new2/20150429new2.csv')
# all_columns = [f for f in real.columns if f in ['stationid','hour','minute','inNums','outNums']]
# X_data = real[all_columns]
# test = test.merge(X_data, on=['stationid','hour','minute'], how='left')
# test.fillna(0, inplace=True)
# print(test[0:30])
# print(test.columns)
# test.to_csv('_10mins/test.csv')

#


# from datetime import datetime
# date = [datetime.strftime(x, '%Y%m%d') for x in list(pd.date_range('20150427', '20150430'))]
# for i in date:
#     bfn='_10mins/'+str(i)+'new.csv'
#     test_sh=pd.read_csv(bfn)
#     test_sh_base=get_base_features(test_sh)
#     del test_sh
#     fn = str(i) + 'new2.csv'
#     path = os.path.join("_10mins/new2", fn)
#     test_sh_base.to_csv(path)
#     # test_sh['date']=test_sh['date'].apply(transfer_datetoday)
#     # test_sh['week']=test_sh['date'].apply(transfer_realweek)
#     # test_sh['weekend'] = test_sh['date'].apply(transfer_realweekend)
#     # test_sh.rename(columns={'cardnums_station_hout_minute':'cardnums_station_hour_minute'}, inplace = True)
#     # test_sh.drop(['Unnamed: 0'], axis=1, inplace=True)
#     # test_sh2=test_sh.copy()
#     # test_sh2.drop(['cardnums_station', 'cardnums_station_hour','cardnums_station_hour_minute','day_since_first'], axis=1, inplace=True)
#     # fn = str(i)+'new1.csv'
#     # fn2=str(i)+'new2.csv'
#     # path = os.path.join("_10mins", fn)
#     # test_sh.to_csv(path)
#     # path = os.path.join("_10mins", fn2)
#     # test_sh2.to_csv(path)